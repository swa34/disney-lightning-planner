import datetime
import io
import json
import os
import tempfile

import openpyxl
from flask import (Flask, flash, make_response, redirect, render_template,
                   request, session, url_for)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from disney_planner import DisneyLightningLanePlanner

app = Flask(__name__)
app.secret_key = os.urandom(24)


# For the index route - no changes needed
@app.route("/")
def index():
    # Clear session data when starting fresh
    session.clear()
    return render_template("index.html")


# For the user_info route - add current_step=1
@app.route("/user_info", methods=["GET", "POST"])
def user_info():
    if request.method == "POST":
        # Initialize planner
        planner = DisneyLightningLanePlanner()

        # Process form data for number of people
        planner.user_data["adults"] = int(request.form.get("adults", 0))
        planner.user_data["children"] = int(request.form.get("children", 0))
        planner.user_data["number_of_people"] = (
            planner.user_data["adults"] + planner.user_data["children"]
        )

        # Process Lightning Lane preferences
        planner.user_data["include_single_pass"] = (
            request.form.get("include_single_pass") == "yes"
        )
        planner.user_data["include_premier"] = (
            request.form.get("include_premier") == "yes"
        )

        # Process resort and park hopping info
        planner.user_data["resort_guest"] = request.form.get("resort_guest") == "yes"
        checkin_date_str = request.form.get("checkin_date")
        if checkin_date_str:
            planner.user_data["checkin_date"] = (
                checkin_date_str  # Store as string, will convert later
            )
        planner.user_data["park_hopping"] = request.form.get("park_hopping") == "yes"

        # Store in session
        session["user_data"] = planner.user_data

        # Redirect to travel dates page
        return redirect(url_for("travel_dates"))

    return render_template("user_info.html", current_step=1)


@app.route("/travel_dates", methods=["GET", "POST"])
def travel_dates():
    if request.method == "POST":
        # Get planner from session
        planner = DisneyLightningLanePlanner()
        planner.user_data = session.get("user_data", {})

        # Process travel dates
        month = int(request.form.get("month"))
        year = int(request.form.get("year"))
        start_day = int(request.form.get("start_day"))
        num_days = int(request.form.get("num_days"))

        # Generate dates
        travel_dates = []
        current_date = datetime.date(year, month, start_day)
        for _ in range(num_days):
            travel_dates.append(current_date.isoformat())
            current_date += datetime.timedelta(days=1)

        planner.user_data["travel_dates"] = travel_dates

        # Store in session
        session["user_data"] = planner.user_data

        return redirect(url_for("park_selection"))

    # Get user data from session for pre-filling the form
    user_data = session.get("user_data", {})

    # Default to today's date
    today = datetime.date.today()
    default_month = today.month
    default_day = today.day
    default_year = today.year

    # If user is a resort guest and provided a check-in date, use that
    if user_data.get("resort_guest") and user_data.get("checkin_date"):
        try:
            checkin_date = datetime.date.fromisoformat(user_data["checkin_date"])
            default_month = checkin_date.month
            default_day = checkin_date.day
            default_year = checkin_date.year
        except (ValueError, TypeError):
            # Fall back to today's date if there's an error
            pass

    return render_template(
        "travel_dates.html",
        current_step=2,
        default_month=default_month,
        default_day=default_day,
        default_year=default_year,
    )


# For the park_selection route - add current_step=3
@app.route("/park_selection", methods=["GET", "POST"])
def park_selection():
    if request.method == "POST":
        # Get planner from session
        planner = DisneyLightningLanePlanner()
        planner.user_data = session.get("user_data", {})

        # Process park selections
        parks = []
        for i in range(len(planner.user_data["travel_dates"])):
            park = request.form.get(f"park_{i}")
            parks.append(park)

        planner.user_data["parks"] = parks

        # Initialize selected_single_passes dictionary
        planner.user_data["selected_single_passes"] = {}
        for park in planner.user_data["parks"]:
            planner.user_data["selected_single_passes"][park] = []

        # Store in session
        session["user_data"] = planner.user_data

        # If user wants Single Pass, redirect to attraction selection page
        if planner.user_data["include_single_pass"]:
            return redirect(url_for("single_pass_selection"))
        else:
            return redirect(url_for("results"))

    # Parse dates from session for template
    user_data = session.get("user_data", {})
    dates = []

    for date_str in user_data.get("travel_dates", []):
        date = datetime.date.fromisoformat(date_str)
        dates.append(
            {"date_str": date_str, "formatted": date.strftime("%A, %B %d, %Y")}
        )

    return render_template("park_selection.html", dates=dates, current_step=3)


# For the single_pass_selection route - add current_step=4s
@app.route("/single_pass_selection", methods=["GET", "POST"])
def single_pass_selection():
    # Get planner from session
    planner = DisneyLightningLanePlanner()
    planner.user_data = session.get("user_data", {})

    if request.method == "POST":
        # Process selected single passes
        for i, park in enumerate(planner.user_data["parks"]):
            selected_attractions = request.form.getlist(f"attractions_{i}")
            planner.user_data["selected_single_passes"][park] = selected_attractions

        # Store in session
        session["user_data"] = planner.user_data

        return redirect(url_for("results"))

    # Prepare park and date information for template
    parks_info = []
    for i, park in enumerate(planner.user_data["parks"]):
        date = datetime.date.fromisoformat(planner.user_data["travel_dates"][i])

        # Get current wait times from API
        wait_times = planner.get_wait_times(park)

        attractions = []
        for attraction, cost in planner.single_pass_costs[park].items():
            # Match attraction name to wait times (simple matching logic)
            wait_time = "Unknown"
            is_open = False
            
            # Try exact match
            if attraction in wait_times:
                wait_time = wait_times[attraction].get("wait_time", "Unknown")
                is_open = wait_times[attraction].get("is_open", False)
            else:
                # Try to find a close match if exact match isn't found
                for wait_time_name, wait_data in wait_times.items():
                    if attraction in wait_time_name or wait_time_name in attraction:
                        wait_time = wait_data.get("wait_time", "Unknown")
                        is_open = wait_data.get("is_open", False)
                        break

            attractions.append({
                "name": attraction,
                "min_cost": cost["min"],
                "max_cost": cost["max"],
                "wait_time": wait_time,
                "is_open": is_open,
            })

        parks_info.append({
            "index": i,
            "name": park,
            "date": date.strftime("%A, %B %d, %Y"),
            "attractions": attractions,
        })

    return render_template("single_pass_selection.html", parks_info=parks_info, current_step=4)

# Update the results route to fix wait_time handling
@app.route("/results")
def results():
    # Get planner from session
    planner = DisneyLightningLanePlanner()
    planner.user_data = session.get("user_data", {})
    
    # Check if we have the necessary data
    if "travel_dates" not in planner.user_data or "parks" not in planner.user_data:
        # Redirect to the start of the process with an error message
        flash("Missing travel information. Please start the planning process again.", "error")
        return redirect(url_for("index"))
        
    # Convert ISO format dates back to datetime objects
    travel_dates = []
    for date_str in planner.user_data["travel_dates"]:
        travel_dates.append(datetime.date.fromisoformat(date_str))
    planner.user_data["travel_dates"] = travel_dates

    # Calculate costs
    scenarios = planner.calculate_costs()

    # Convert datetime objects to strings for JSON serialization
    for scenario_key, scenario in scenarios.items():
        for i, daily in enumerate(scenario["daily"]):
            scenario["daily"][i]["date"] = daily["date"].strftime("%A, %B %d, %Y")

    # Get park recommendations and tips
    park_recommendations = {}

    for i, park in enumerate(planner.user_data["parks"]):
        date = planner.user_data["travel_dates"][i]

        # Get recommended attractions
        if park == "Animal Kingdom":
            recommended = {
                "tiers": False,
                "attractions": planner.recommended_attractions[park]["Recommended"],
            }
        else:
            recommended = {
                "tiers": True,
                "tier1": planner.recommended_attractions[park]["Tier 1"],
                "tier2": planner.recommended_attractions[park]["Tier 2"],
            }

        # Get single pass information
        single_passes = []
        if planner.user_data.get("include_single_pass", False):
            selected = planner.user_data.get("selected_single_passes", {}).get(park, [])

            for attraction in selected:
                cost_range = planner.single_pass_costs[park][attraction]
                single_passes.append({
                    "name": attraction,
                    "min": cost_range["min"],
                    "max": cost_range["max"],
                })

        # Get premier pass info
        premier_info = None
        if planner.user_data.get("include_premier", False):
            premier_cost = planner.premier_pass_costs[park]
            premier_info = {"min": premier_cost["min"], "max": premier_cost["max"]}

        # Get tips
        tips = planner.park_tips[park]

        # Create park recommendation
        park_recommendations[park] = {
            "date": date.strftime("%A, %B %d, %Y"),
            "recommended": recommended,
            "single_passes": single_passes,
            "premier_info": premier_info,
            "tips": tips,
        }
    
    # Safely get booking information
    try:
        booking_dates = planner.get_booking_dates()
        booking_date = planner.get_booking_date()
        milestones = planner.get_trip_milestones()
    except Exception as e:
        # If there's an error getting booking dates, set them to None
        booking_dates = []
        booking_date = None
        milestones = {}
        # Log the error for debugging
        print(f"Error getting booking dates: {e}")

    # Store calculated data in session for PDF and Excel downloads
    session['scenarios'] = scenarios
    session['park_recommendations'] = park_recommendations
    session['booking_dates'] = [date.isoformat() if date else None for date in booking_dates]
    session['booking_date'] = booking_date.isoformat() if booking_date else None
    session['general_tips'] = planner.general_tips
    
    # For milestones, we need to handle the date objects for serialization
    serialized_milestones = {}
    for key, value in milestones.items():
        if isinstance(value, datetime.date):
            serialized_milestones[key] = value.isoformat()
        elif isinstance(value, list):
            serialized_milestones[key] = [date.isoformat() for date in value]
        else:
            serialized_milestones[key] = value
    
    session['milestones'] = serialized_milestones

    # Add attribution for Queue-Times.com
    queue_times_attribution = "Wait time data powered by Queue-Times.com"

    # Pass the zip function to the template
    return render_template(
        "results.html",
        user_data=planner.user_data,
        scenarios=scenarios,
        park_recommendations=park_recommendations,
        general_tips=planner.general_tips,
        zip=zip,
        booking_dates=booking_dates,
        booking_date=booking_date,
        milestones=milestones,
        current_date=datetime.date.today(),
        current_step=5,
        queue_times_attribution=queue_times_attribution,
    )


@app.route("/restaurants")
def restaurant_finder():
    """Page for the restaurant finder tool"""
    return render_template("restaurant_finder.html")


@app.route("/dining-calculator")
def dining_calculator():
    """Page for the dining cost calculator tool"""
    return render_template("restaurant_calculator.html")


@app.route("/contact")
def contact():
    """Page for the contact information"""
    return render_template("contact.html")

@app.route("/download_ics")
def download_ics():
    label = request.args.get("label", "Disney World Milestone")
    date_str = request.args.get("date")

    try:
        date = datetime.date.fromisoformat(date_str)
    except Exception:
        return "Invalid date", 400

    ics_content = f"""BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//WDWKnow//Milestone Planner//EN
BEGIN:VEVENT
SUMMARY:{label}
DTSTART;VALUE=DATE:{date.strftime('%Y%m%d')}
DTEND;VALUE=DATE:{(date + datetime.timedelta(days=1)).strftime('%Y%m%d')}
DESCRIPTION:Disney World planning milestone - time to book Lightning Lane passes!
END:VEVENT
END:VCALENDAR"""

    response = make_response(ics_content)
    response.headers.set("Content-Type", "text/calendar")
    response.headers.set("Content-Disposition", f"attachment; filename=\"{label}.ics\"")
    return response

def parse_date(date_str):
    """Parse a date string in various formats to a datetime.date object."""
    if not date_str:
        return None
        
    try:
        # Try ISO format first
        return datetime.date.fromisoformat(date_str)
    except ValueError:
        pass
        
    try:
        # Try standard datetime format
        return datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        pass
    
    try:
        # Try the GMT format: 'Wed, 30 Apr 2025 00:00:00 GMT'
        return datetime.datetime.strptime(date_str, '%a, %d %b %Y %H:%M:%S GMT').date()
    except ValueError:
        pass
    
    try:
        # Try another common format
        return datetime.datetime.strptime(date_str, '%a, %b %d, %Y').date()
    except ValueError:
        pass
    
    # If all else fails, return None
    print(f"Could not parse date: {date_str}")
    return None

# Complete download_pdf function with fixed booking dates for off-site guests
@app.route("/download_pdf")
def download_pdf():
    """Generate and download a PDF of the plan results using ReportLab"""
    # Get planner data from session
    user_data = session.get("user_data", {})
    scenarios = session.get("scenarios", {})
    park_recommendations = session.get("park_recommendations", {})
    general_tips = session.get("general_tips", [])
    
    # Get booking information
    booking_dates_str = session.get("booking_dates", [])
    booking_date_str = session.get("booking_date")
    milestones_str = session.get("milestones", {})
    
    # Convert dates back from various string formats to datetime objects
    booking_dates = []
    for date_str in booking_dates_str:
        if date_str:  # Only process non-None values
            booking_dates.append(parse_date(date_str))
        else:
            booking_dates.append(None)
    
    booking_date = parse_date(booking_date_str) if booking_date_str else None
    
    # Convert milestones dates
    milestones = {}
    for key, value in milestones_str.items():
        if isinstance(value, str):
            milestones[key] = parse_date(value)
        elif isinstance(value, list):
            milestones[key] = [parse_date(date_str) if date_str else None for date_str in value]
        else:
            milestones[key] = value
    
    # Make sure travel_dates are datetime objects
    if "travel_dates" in user_data and user_data["travel_dates"]:
        travel_dates = []
        for date_val in user_data["travel_dates"]:
            if isinstance(date_val, str):
                parsed_date = parse_date(date_val)
                if parsed_date:
                    travel_dates.append(parsed_date)
            elif isinstance(date_val, dict) and "date" in date_val:
                # Handle the case where dates are stored as objects
                parsed_date = parse_date(date_val["date"])
                if parsed_date:
                    travel_dates.append(parsed_date)
            elif hasattr(date_val, 'strftime'):  # It's already a date/datetime object
                if isinstance(date_val, datetime.datetime):
                    travel_dates.append(date_val.date())
                else:
                    travel_dates.append(date_val)
        
        if travel_dates:  # Only update if we successfully parsed dates
            user_data["travel_dates"] = travel_dates
    
    # Create a PDF using ReportLab
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    # Create PDF document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Create custom styles - use unique names that aren't in the default stylesheet
    styles.add(ParagraphStyle(name='DisneyTitle', 
                             parent=styles['Heading1'], 
                             fontName='Helvetica-Bold',
                             fontSize=18, 
                             textColor=colors.HexColor('#0078D2'),
                             spaceAfter=12))
    
    styles.add(ParagraphStyle(name='DisneyHeading2', 
                             parent=styles['Heading2'], 
                             fontName='Helvetica-Bold',
                             fontSize=14, 
                             textColor=colors.HexColor('#0078D2'),
                             spaceBefore=12,
                             spaceAfter=6))
    
    styles.add(ParagraphStyle(name='DisneyHeading3', 
                             parent=styles['Heading3'], 
                             fontName='Helvetica-Bold',
                             fontSize=12, 
                             textColor=colors.HexColor('#6F00FF'),
                             spaceBefore=10,
                             spaceAfter=6))
    
    # Collection of elements for the PDF
    elements = []
    
    # Title
    elements.append(Paragraph("Disney World Lightning Lane Plan", styles["DisneyTitle"]))
    elements.append(Paragraph(f"Generated on {datetime.date.today().strftime('%A, %B %d, %Y')}", styles["Normal"]))
    elements.append(Spacer(1, 0.25*inch))
    
    # Trip Summary
    elements.append(Paragraph("Trip Summary", styles["DisneyHeading2"]))
    
    # Basic info
    data = [
        ["Party Size:", f"{user_data.get('adults', 0)} adults" + (f", {user_data.get('children', 0)} children" if user_data.get('children', 0) > 0 else "")],
        ["Resort Guest:", "Yes" if user_data.get('resort_guest', False) else "No"],
        ["Park Hopper:", "Yes" if user_data.get('park_hopping', False) else "No"]
    ]
    
    # Create table
    t = Table(data, colWidths=[1.5*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.2*inch))
    
    # Park Visit Schedule
    if "travel_dates" in user_data and user_data["travel_dates"] and "parks" in user_data:
        elements.append(Paragraph("Park Visit Schedule", styles["DisneyHeading3"]))
        
        visit_data = [["Date", "Park"]]
        for i, (date, park) in enumerate(zip(user_data.get('travel_dates', []), user_data.get('parks', []))):
            if date and hasattr(date, 'strftime'):  # Make sure date is a valid date object
                visit_data.append([date.strftime('%A, %B %d, %Y'), park])
        
        if len(visit_data) > 1:  # Only add table if there are rows other than header
            # Create table for park visits
            t = Table(visit_data, colWidths=[2.5*inch, 3*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.2*inch))
    
    # Lightning Lane Booking Information
    elements.append(Paragraph("Lightning Lane Booking Information", styles["DisneyHeading3"]))
    
    if user_data.get('resort_guest', False) and booking_date and hasattr(booking_date, 'strftime'):
        elements.append(Paragraph("As a Disney Resort guest, you can book Lightning Lane passes for your entire stay starting at 7:00 AM ET on:", styles["Normal"]))
        elements.append(Paragraph(booking_date.strftime('%A, %B %d, %Y'), styles["Normal"]))
    elif user_data.get('travel_dates') and len(user_data.get('travel_dates', [])) > 0:
        elements.append(Paragraph("As an off-site guest, you can book Lightning Lane passes starting 3 days before each park visit:", styles["Normal"]))
        
        booking_data = [["Visit Date", "Park", "Book On"]]
        
        # Check if we have valid booking dates
        has_valid_booking_data = False
        for i, (visit_date, park) in enumerate(zip(user_data.get('travel_dates', []), user_data.get('parks', []))):
            if hasattr(visit_date, 'strftime'):
                book_date = None
                # Try to get the booking date from the list
                if i < len(booking_dates) and booking_dates[i] and hasattr(booking_dates[i], 'strftime'):
                    book_date = booking_dates[i]
                else:
                    # Fall back to calculating it
                    book_date = visit_date - datetime.timedelta(days=3)
                
                booking_data.append([
                    visit_date.strftime('%A, %B %d'), 
                    park, 
                    book_date.strftime('%A, %B %d')
                ])
                has_valid_booking_data = True
        
        if has_valid_booking_data:  # Only add table if there are valid rows
            t = Table(booking_data, colWidths=[2*inch, 2*inch, 2*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            ]))
            elements.append(t)
    
    # Trip Planning Milestones
    if milestones:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("Trip Planning Milestones", styles["DisneyHeading3"]))
        
        milestone_data = [["Milestone", "Date"]]
        
        # Process milestones with single dates
        for label, date_value in milestones.items():
            if isinstance(date_value, (datetime.date, datetime.datetime)) and hasattr(date_value, 'strftime'):
                milestone_data.append([label, date_value.strftime('%A, %B %d, %Y')])
            
        # Add the table if there are any milestone dates
        if len(milestone_data) > 1:
            t = Table(milestone_data, colWidths=[3*inch, 3*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            ]))
            elements.append(t)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Page break
    elements.append(Paragraph("Lightning Lane Options", styles["DisneyHeading2"]))
    
    # Scenarios
    for scenario_key, scenario in scenarios.items():
        elements.append(Paragraph(scenario['name'], styles["DisneyHeading3"]))
        
        scenario_data = [["Date", "Park", "Cost Range", "Single Passes"]]
        
        for daily in scenario['daily']:
            date_parts = daily['date'].split(',')[0] if isinstance(daily['date'], str) else "Unknown Date"
            single_passes = ', '.join(daily.get('single_passes', [])) if daily.get('single_passes') else "None"
            
            scenario_data.append([
                date_parts,
                daily['park'],
                f"${daily['min']} - ${daily['max']}",
                single_passes
            ])
        
        # Add total row
        scenario_data.append([
            "Total Cost:", 
            "", 
            f"${scenario['total_min']} - ${scenario['total_max']}", 
            ""
        ])
        
        # Create the table
        col_widths = [1.2*inch, 1.5*inch, 1.5*inch, 2.3*inch]
        t = Table(scenario_data, colWidths=col_widths)
        
        # Style the table
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            
            # Style the total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (0, -1), 'Helvetica-Bold'),
        ]
        
        t.setStyle(TableStyle(table_style))
        elements.append(t)
        
        # Add the recommendation
        elements.append(Spacer(1, 0.15*inch))
        recommendation = ""
        if scenario_key == "scenario1":
            recommendation = "This option provides comprehensive coverage with your selected Lightning Lane Single Passes combined with Multi Pass for most attractions."
        elif scenario_key == "scenario2":
            recommendation = "Best for budget-conscious visitors who are willing to use standby lines for the most popular attractions or visit during low-crowd periods."
        elif scenario_key == "scenario3":
            recommendation = "Premier Pass is very expensive but offers maximum convenience. Consider this option for just one day of your trip at your highest-priority park, rather than for every day."
        elif scenario_key == "scenario4":
            if user_data.get('include_single_pass', False):
                recommendation = "This mixed approach offers the best balance of cost and experience. It prioritizes Lightning Lane passes where they save you the most time and skips them where they're less critical."
            else:
                recommendation = "This approach focuses on Multi Pass value while avoiding the additional cost of Single Passes. You'll need to use rope drop strategies or standby lines for the premium attractions."
        
        elements.append(Paragraph(f"Recommendation: {recommendation}", styles["Normal"]))
        elements.append(Spacer(1, 0.25*inch))
    
    # Page break
    elements.append(Paragraph("Park-by-Park Recommendations", styles["DisneyHeading2"]))
    
    # Park Recommendations
    for park, park_data in park_recommendations.items():
        # Park header with date
        park_date = park_data.get('date', '')
        elements.append(Paragraph(f"{park} - {park_date}", styles["DisneyHeading3"]))
        
        # Recommended Lightning Lane Selections
        elements.append(Paragraph("Recommended Lightning Lane Selections:", styles["Normal"]))
        
        if park_data.get('recommended', {}).get('tiers', False):
            # Tier 1
            elements.append(Paragraph("Tier 1 (Choose 1):", ParagraphStyle(
                "BulletHeading", 
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                leftIndent=10,
                spaceBefore=6
            )))
            
            for attraction in park_data.get('recommended', {}).get('tier1', []):
                elements.append(Paragraph(f"• {attraction}", ParagraphStyle(
                    "Bullet", 
                    parent=styles["Normal"],
                    leftIndent=20
                )))
            
            # Tier 2
            elements.append(Paragraph("Tier 2 (Choose 2):", ParagraphStyle(
                "BulletHeading", 
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                leftIndent=10,
                spaceBefore=6
            )))
            
            for attraction in park_data.get('recommended', {}).get('tier2', []):
                elements.append(Paragraph(f"• {attraction}", ParagraphStyle(
                    "Bullet", 
                    parent=styles["Normal"],
                    leftIndent=20
                )))
        else:
            # No tiers
            elements.append(Paragraph("Recommended Attractions:", ParagraphStyle(
                "BulletHeading", 
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                leftIndent=10,
                spaceBefore=6
            )))
            
            for attraction in park_data.get('recommended', {}).get('attractions', []):
                elements.append(Paragraph(f"• {attraction}", ParagraphStyle(
                    "Bullet", 
                    parent=styles["Normal"],
                    leftIndent=20
                )))
        
        # Single passes
        if user_data.get('include_single_pass', False) and park_data.get('single_passes', []):
            elements.append(Paragraph("Your Selected Single Passes:", ParagraphStyle(
                "BulletHeading", 
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                spaceBefore=10
            )))
            
            for pass_info in park_data.get('single_passes', []):
                elements.append(Paragraph(f"• {pass_info.get('name', '')} (${pass_info.get('min', 0)}-${pass_info.get('max', 0)} per person)", ParagraphStyle(
                    "Bullet", 
                    parent=styles["Normal"],
                    leftIndent=20
                )))
        
        # Tips
        elements.append(Paragraph(f"Tips for {park}:", ParagraphStyle(
            "BulletHeading", 
            parent=styles["Normal"],
            fontName="Helvetica-Bold", 
            spaceBefore=10
        )))
        
        for tip in park_data.get('tips', []):
            elements.append(Paragraph(f"• {tip}", ParagraphStyle(
                "Bullet", 
                parent=styles["Normal"],
                leftIndent=20
            )))
        
        elements.append(Spacer(1, 0.25*inch))
    
    # General Tips
    elements.append(Paragraph("General Lightning Lane Tips", styles["DisneyHeading2"]))
    
    # Single Pass Strategy
    elements.append(Paragraph("Single Pass Strategy:", ParagraphStyle(
        "BulletHeading", 
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        spaceBefore=6
    )))
    
    for tip in ["Book as early in the day as possible for popular attractions", 
               "Purchase early as these often sell out in advance", 
               "Remember you can purchase up to 2 Single Passes per day across all parks"]:
        elements.append(Paragraph(f"• {tip}", ParagraphStyle(
            "Bullet", 
            parent=styles["Normal"],
            leftIndent=20
        )))
    
    # Multi Pass Strategy
    elements.append(Paragraph("Multi Pass Strategy:", ParagraphStyle(
        "BulletHeading", 
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        spaceBefore=10
    )))
    
    for tip in ["Pre-book your 3 attractions before your trip", 
               "Use tier rules: 1 Tier 1 + 2 Tier 2 attractions (except Animal Kingdom)", 
               "After using your first Lightning Lane, immediately book your next one", 
               "You can modify existing bookings without canceling them"]:
        elements.append(Paragraph(f"• {tip}", ParagraphStyle(
            "Bullet", 
            parent=styles["Normal"],
            leftIndent=20
        )))
    
    # Additional Tips
    elements.append(Paragraph("Additional Tips:", ParagraphStyle(
        "BulletHeading", 
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        spaceBefore=10
    )))
    
    for tip in general_tips:
        elements.append(Paragraph(f"• {tip}", ParagraphStyle(
            "Bullet", 
            parent=styles["Normal"],
            leftIndent=20
        )))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("Note: This is an unofficial planning tool and is not affiliated with Walt Disney World Resort or the Walt Disney Company.", 
                             ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey)))
    elements.append(Paragraph("All prices shown are estimates based on historical and seasonal data. Actual costs may vary.", 
                             ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey)))
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value from the BytesIO buffer and create the response
    pdf_value = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = make_response(pdf_value)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=Disney_Lightning_Lane_Plan.pdf'
    
    return response

# Complete download_excel function with fixed booking dates for off-site guests
@app.route("/download_excel")
def download_excel():
    """Generate and download an Excel file of the plan results"""
    # Get planner data from session
    user_data = session.get("user_data", {})
    scenarios = session.get("scenarios", {})
    park_recommendations = session.get("park_recommendations", {})
    general_tips = session.get("general_tips", [])
    
    # Get booking information
    booking_dates_str = session.get("booking_dates", [])
    booking_date_str = session.get("booking_date")
    milestones_str = session.get("milestones", {})
    
    # Convert dates back from various string formats to datetime objects
    booking_dates = []
    for date_str in booking_dates_str:
        if date_str:  # Only process non-None values
            booking_dates.append(parse_date(date_str))
        else:
            booking_dates.append(None)
    
    booking_date = parse_date(booking_date_str) if booking_date_str else None
    
    # Convert milestones dates
    milestones = {}
    for key, value in milestones_str.items():
        if isinstance(value, str):
            milestones[key] = parse_date(value)
        elif isinstance(value, list):
            milestones[key] = [parse_date(date_str) if date_str else None for date_str in value]
        else:
            milestones[key] = value
    
    # Make sure travel_dates are datetime objects
    if "travel_dates" in user_data and user_data["travel_dates"]:
        travel_dates = []
        for date_val in user_data["travel_dates"]:
            if isinstance(date_val, str):
                parsed_date = parse_date(date_val)
                if parsed_date:
                    travel_dates.append(parsed_date)
            elif isinstance(date_val, dict) and "date" in date_val:
                # Handle the case where dates are stored as objects
                parsed_date = parse_date(date_val["date"])
                if parsed_date:
                    travel_dates.append(parsed_date)
            elif hasattr(date_val, 'strftime'):  # It's already a date/datetime object
                if isinstance(date_val, datetime.datetime):
                    travel_dates.append(date_val.date())
                else:
                    travel_dates.append(date_val)
        
        if travel_dates:  # Only update if we successfully parsed dates
            user_data["travel_dates"] = travel_dates
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Create styles
    title_font = Font(name='Arial', size=14, bold=True, color='0078D2')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0078D2', end_color='0078D2', fill_type='solid')
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Trip Summary Sheet
    ws = wb.active
    ws.title = "Trip Summary"
    
    # Add title
    ws['A1'] = "Disney World Lightning Lane Plan"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Trip details
    ws['A3'] = "Trip Details"
    ws['A3'].font = subheader_font
    ws.merge_cells('A3:B3')
    
    ws['A4'] = "Party Size:"
    ws['B4'] = f"{user_data.get('adults', 0)} adults" + (f", {user_data.get('children', 0)} children" if user_data.get('children', 0) > 0 else "")
    
    ws['A5'] = "Resort Guest:"
    ws['B5'] = "Yes" if user_data.get('resort_guest', False) else "No"
    
    ws['A6'] = "Park Hopper:"
    ws['B6'] = "Yes" if user_data.get('park_hopping', False) else "No"
    
    # Park Visit Schedule
    ws['A8'] = "Park Visit Schedule"
    ws['A8'].font = subheader_font
    ws.merge_cells('A8:C8')
    
    ws['A9'] = "Date"
    ws['B9'] = "Day"
    ws['C9'] = "Park"
    
    for cell in ws['A9:C9'][0]:
        cell.font = header_font
        cell.fill = header_fill
    
    row = 10
    if "travel_dates" in user_data and "parks" in user_data:
        for i, (date, park) in enumerate(zip(user_data.get('travel_dates', []), user_data.get('parks', []))):
            if date and hasattr(date, 'strftime'):  # Make sure date is a valid date object
                ws[f'A{row}'] = date
                ws[f'B{row}'] = date.strftime('%A')
                ws[f'C{row}'] = park
                row += 1
    
    # Booking dates
    row += 2
    ws[f'A{row}'] = "Lightning Lane Booking Information"
    ws[f'A{row}'].font = subheader_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    if user_data.get('resort_guest', False) and booking_date and hasattr(booking_date, 'strftime'):
        ws[f'A{row}'] = "As a Disney Resort guest, you can book Lightning Lane passes for your entire stay starting at 7:00 AM ET on:"
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        ws[f'A{row}'] = booking_date
        ws[f'B{row}'] = booking_date.strftime('%A')
        row += 2
    elif user_data.get('travel_dates') and len(user_data.get('travel_dates', [])) > 0:
        ws[f'A{row}'] = "As an off-site guest, you can book Lightning Lane passes starting 3 days before each park visit:"
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws[f'A{row}'] = "Visit Date"
        ws[f'B{row}'] = "Park"
        ws[f'C{row}'] = "Book On"
        
        for cell in ws[f'A{row}:C{row}'][0]:
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        for i, (visit_date, park) in enumerate(zip(user_data.get('travel_dates', []), user_data.get('parks', []))):
            if hasattr(visit_date, 'strftime'):
                book_date = None
                # Try to get the booking date from the list
                if i < len(booking_dates) and booking_dates[i] and hasattr(booking_dates[i], 'strftime'):
                    book_date = booking_dates[i]
                else:
                    # Fall back to calculating it
                    book_date = visit_date - datetime.timedelta(days=3)
                
                ws[f'A{row}'] = visit_date
                ws[f'B{row}'] = park
                ws[f'C{row}'] = book_date
                row += 1
    
    # Add milestones section
    if milestones:
        row += 2
        ws[f'A{row}'] = "Trip Planning Milestones"
        ws[f'A{row}'].font = subheader_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Header row
        ws[f'A{row}'] = "Milestone"
        ws[f'B{row}'] = "Date"
        
        for cell in ws[f'A{row}:B{row}'][0]:
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        # Process milestones with single dates
        for label, date_value in milestones.items():
            if isinstance(date_value, (datetime.date, datetime.datetime)) and hasattr(date_value, 'strftime'):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = date_value
                row += 1
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create Lightning Lane Options sheet
    ws = wb.create_sheet(title="Lightning Lane Options")
    
    ws['A1'] = "Lightning Lane Options"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    for scenario_key, scenario in scenarios.items():
        ws[f'A{row}'] = scenario['name']
        ws[f'A{row}'].font = subheader_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = "Park"
        ws[f'C{row}'] = "Cost Range"
        ws[f'D{row}'] = "Single Passes"
        
        for cell in ws[f'A{row}:D{row}'][0]:
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        for daily in scenario['daily']:
            if isinstance(daily['date'], str):
                date_parts = daily['date'].split(',')[0]  # Get just the day part, e.g. "Monday, April 15" -> "Monday"
            else:
                date_parts = "Unknown Date"
                
            ws[f'A{row}'] = date_parts
            ws[f'B{row}'] = daily['park']
            ws[f'C{row}'] = f"${daily['min']} - ${daily['max']}"
            
            if 'single_passes' in daily and daily['single_passes']:
                ws[f'D{row}'] = ', '.join(daily['single_passes'])
            else:
                ws[f'D{row}'] = "None"
            
            row += 1
        
        ws[f'A{row}'] = "Total Cost:"
        ws[f'C{row}'] = f"${scenario['total_min']} - ${scenario['total_max']}"
        ws[f'A{row}'].font = subheader_font
        ws[f'C{row}'].font = subheader_font
        
        row += 3  # Space between scenarios
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create Park Recommendations sheet
    ws = wb.create_sheet(title="Park Recommendations")
    
    ws['A1'] = "Park-by-Park Recommendations"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    for park, park_data in park_recommendations.items():
        ws[f'A{row}'] = f"{park} - {park_data.get('date', '')}"
        ws[f'A{row}'].font = subheader_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws[f'A{row}'] = "Recommended Lightning Lane Selections:"
        ws[f'A{row}'].font = subheader_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        if park_data.get('recommended', {}).get('tiers', False):
            ws[f'A{row}'] = "Tier 1 (Choose 1):"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for attraction in park_data.get('recommended', {}).get('tier1', []):
                ws[f'A{row}'] = "• " + attraction
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
            
            row += 1
            ws[f'A{row}'] = "Tier 2 (Choose 2):"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for attraction in park_data.get('recommended', {}).get('tier2', []):
                ws[f'A{row}'] = "• " + attraction
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
        else:
            ws[f'A{row}'] = "Recommended Attractions:"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for attraction in park_data.get('recommended', {}).get('attractions', []):
                ws[f'A{row}'] = "• " + attraction
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
        
        # Single passes
        if user_data.get('include_single_pass', False) and park_data.get('single_passes', []):
            row += 1
            ws[f'A{row}'] = "Selected Single Passes:"
            ws[f'A{row}'].font = subheader_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            for pass_info in park_data.get('single_passes', []):
                ws[f'A{row}'] = f"• {pass_info.get('name', '')} (${pass_info.get('min', 0)}-${pass_info.get('max', 0)} per person)"
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
        
        # Tips
        row += 1
        ws[f'A{row}'] = "Tips for " + park + ":"
        ws[f'A{row}'].font = subheader_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for tip in park_data.get('tips', []):
            ws[f'A{row}'] = "• " + tip
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        row += 3  # Space between parks
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create General Tips sheet
    ws = wb.create_sheet(title="General Tips")
    
    ws['A1'] = "General Lightning Lane Tips"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    ws[f'A{row}'] = "Single Pass Strategy:"
    ws[f'A{row}'].font = subheader_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    tips = [
        "Book as early in the day as possible for popular attractions",
        "Purchase early as these often sell out in advance",
        "Remember you can purchase up to 2 Single Passes per day across all parks"
    ]
    
    for tip in tips:
        ws[f'A{row}'] = "• " + tip
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Multi Pass Strategy:"
    ws[f'A{row}'].font = subheader_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    tips = [
        "Pre-book your 3 attractions before your trip",
        "Use tier rules: 1 Tier 1 + 2 Tier 2 attractions (except Animal Kingdom)",
        "After using your first Lightning Lane, immediately book your next one",
        "You can modify existing bookings without canceling them"
    ]
    
    for tip in tips:
        ws[f'A{row}'] = "• " + tip
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Additional Tips:"
    ws[f'A{row}'].font = subheader_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for tip in general_tips:
        ws[f'A{row}'] = "• " + tip
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response with Excel file
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Disney_Lightning_Lane_Plan.xlsx'
    
    return response

@app.route("/wait_times")
def wait_times():
    """Page for viewing current wait times by park"""
    # Create planner to access wait time data
    planner = DisneyLightningLanePlanner()
    
    # Dictionary to store wait times for each park
    parks_wait_times = {}
    
    # Fetch wait times for all Disney World parks
    for park_name, park_id in planner.queue_times_park_ids.items():
        parks_wait_times[park_name] = planner.get_wait_times(park_name)
    
    # Add attribution for Queue-Times.com
    queue_times_attribution = "Wait time data powered by Queue-Times.com"
    
    return render_template(
        "wait_times.html", 
        parks_wait_times=parks_wait_times,
        queue_times_attribution=queue_times_attribution
    )


if __name__ == "__main__":
    app.run(debug=True)